import openpyxl as opx

#========================================================================================================
def getStaticDays(filename):
    try:
        static_workbook = openFile(filename + '.xlsx')
        static_sheet = static_workbook[static_sheet_list[0]]
        counter = 12
        print('importing days from static file to payroll format file...')
        while counter < 27:
            static_data = static_sheet.cell(row=counter, column=1).value
            days_list.append(static_data)
            counter += 1
    finally:
        closeFile(static_workbook)
        insertDaysToPayroll()


def insertDaysToPayroll():
    try:
        payroll_workbook = openFile('PAYROLL format.xlsx')
        payroll_sheet = payroll_workbook['sheet3']

        """ FOR DATES AND DAYS """
        print('processing imported data...')
        days_period = 15
        for column_num in date_columns_list:							#starting columns in sheet3
            for row in rows_list:								#starting rows in sheet3
                counter = 0
                while counter <= 15:							#loop to fill the rows per person days
                    counter += 1
                    row_num = row
                    for day in days_list:							#monthday and weekday of payroll coverage
                        payroll_sheet.cell(row=row_num, column=column_num, value=day)
                        row_num += 1
        print('done importing data...')

        """ FOR EMPLOYEE TIMESTAMPS """
        static_workbook = openFile('09Static.xlsx')
        static_sheet_counter = 0
        static_sheet = static_workbook[static_sheet_list[static_sheet_counter]]
        emp_list = payroll_emp_dtr.keys()
        static_col_counter = 0
        row_count = 12

        print('initializing employees\' timestamps...')
        # sheet_counter = 1
        # while sheet_counter <= 3:
        forloop_counter = 1
        for emp in emp_list:
            emp_row = payroll_emp_dtr[emp]['pr_row']
            emp_column_in = payroll_emp_dtr[emp]['pr_in_column']
            emp_column_out = payroll_emp_dtr[emp]['pr_out_column']
            data_column_in = static_column_list_in[static_col_counter]
            data_column_out = static_column_list_out[static_col_counter]

            """for IN timestamps"""
            while row_count <= 27:
                data_static_in = static_sheet.cell(row=row_count, column=data_column_in).value
                payroll_sheet.cell(row=emp_row, column=emp_column_in, value=data_static_in)
                row_count += 1
                emp_row += 1
            row_count = 12
            emp_row = payroll_emp_dtr[emp]['pr_row']   

            """for OUT timestamps"""                
            while row_count <= 27:
                data_static_out = static_sheet.cell(row=row_count, column=data_column_out).value
                payroll_sheet.cell(row=emp_row, column=emp_column_out, value=data_static_out)
                row_count += 1
                emp_row += 1
            row_count = 12
            static_col_counter += 1
            forloop_counter += 1

            if forloop_counter == 4:
                static_sheet_counter += 1
                forloop_counter = 1
                static_col_counter = 0
                static_workbook.remove(static_sheet)
                new_static_sheet = static_workbook[static_sheet_list[static_sheet_counter]]
                static_sheet = new_static_sheet
            else:
                pass

    finally:
        print('done on processing employee\'s timestamps...')   
        payroll_workbook.save('test.xlsx')
        payroll_workbook.close()
        static_workbook.close()    
        print('calculating emloyee\'s working days')
        empWorkingDays()


def empWorkingDays():
    ewd_wb = openFile('test.xlsx')
    ewd_sheet = ewd_wb['sheet3']
    ewd_row_list = rows_list
    ewd_column_list = [3, 10, 17, 24]
    ewd_to_write_row_list = [9, 10, 7, 8, 11, 15, 13, 14, 16, 17, 18, 19, 22, 20, 21]

    workday = 0
    wd_list = []

    for ewd_row in ewd_row_list:
        ewd_row_count = ewd_row
        for ewd_column in ewd_column_list:
            ewd_column_count = ewd_column
            for ewd_data_count in range(16):
                ewd_data1 = ewd_sheet.cell(row=ewd_row_count, column=ewd_column_count).value
                if ewd_data1 == None:
                    pass
                else:
                    workday += 1
                ewd_row_count += 1
            wd_list.append(workday)
            ewd_row_count = ewd_row
            workday = 0

    print('importing of working days to payroll format sheet...')
    payrollformat_sheet = ewd_wb['Payroll Format']
    ewd_data_list_counter = 0
    for ewd_row_to_write in ewd_to_write_row_list:
        payrollformat_sheet.cell(row=ewd_row_to_write, column=10, value=wd_list[ewd_data_list_counter])
        ewd_data_list_counter += 1


    ewd_wb.save('test.xlsx')
    ewd_wb.close()
    print('done importing!')
    input('Press enter to terminate Project Venice. Thank you!')

    
#=======================================================================================================
   
def openFile(filename):
    workbook = opx.load_workbook(filename)
    return workbook
    

def closeFile(filename):
    filename.close()


def start():
    print('HELLO, I\'M VENICE. :D')
    static_file = str(input('Please type the filename of static file: '))
    print(f'opening file {static_file} ...')
    getStaticDays(static_file)
    


#=======================================================================================================

"""for dates"""
rows_list = [7, 30, 52, 74]
date_columns_list = [2, 9, 16, 23]
days_list = []

"for DTR"
static_sheet_list = ['1.2.5', '17.21.30', '31.35.1901', '1902.11801.11802', '11803.11804.11805', '11806.11807.11808', '11809.11810.11812', '11813.11814']
static_column_list_in = [2, 17, 32]
static_column_list_out = [4, 19, 34]
payroll_emp_dtr = {
    'emp001':{'pr_row':52, 'pr_in_column':10, 'pr_out_column':12},
    'emp002':{'pr_row':74, 'pr_in_column':10, 'pr_out_column':12},
    'emp003':{'pr_row':52, 'pr_in_column':24, 'pr_out_column':26},
    'emp004':{'pr_row':30, 'pr_in_column':17, 'pr_out_column':19},
    'emp005':{'pr_row':72, 'pr_in_column':24, 'pr_out_column':26},
    'emp006':{'pr_row':52, 'pr_in_column':3, 'pr_out_column':5},
    'emp007':{'pr_row':74, 'pr_in_column':3, 'pr_out_column':5},
    'emp008':{'pr_row':74, 'pr_in_column':17, 'pr_out_column':19},
    'emp009':{'pr_row':52, 'pr_in_column':17, 'pr_out_column':19},
    'emp010':{'pr_row':30, 'pr_in_column':3, 'pr_out_column':5},
    'emp011':{'pr_row':72, 'pr_in_column':24, 'pr_out_column':26},
    'emp012':{'pr_row':72, 'pr_in_column':24, 'pr_out_column':26},
    'emp013':{'pr_row':72, 'pr_in_column':24, 'pr_out_column':26},
    'emp014':{'pr_row':72, 'pr_in_column':24, 'pr_out_column':26},
    'emp015':{'pr_row':72, 'pr_in_column':24, 'pr_out_column':26},
    'emp016':{'pr_row':7, 'pr_in_column':3, 'pr_out_column':5},
    'emp017':{'pr_row':7, 'pr_in_column':10, 'pr_out_column':12},
    'emp018':{'pr_row':7, 'pr_in_column':17, 'pr_out_column':19},
    'emp019':{'pr_row':7, 'pr_in_column':24, 'pr_out_column':26},
    'emp020':{'pr_row':72, 'pr_in_column':24, 'pr_out_column':26},
    'emp021':{'pr_row':72, 'pr_in_column':24, 'pr_out_column':26},    
    'emp022':{'pr_row':30, 'pr_in_column':10, 'pr_out_column':12},
    'emp023':{'pr_row':30, 'pr_in_column':24, 'pr_out_column':26}
}

start()


